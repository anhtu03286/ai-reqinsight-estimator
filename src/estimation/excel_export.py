"""
4-tab Excel quotation export.
KL-P-002: all cost cells must use formulas referencing rate/effort cells — never hardcoded computed values.
"""
import io
import uuid
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import select
from src.models.wbs_item import WBSItem
from src.models.rate_card import RateCard, RateCardEntry

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center")


def _header(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    return cell


def _border_all(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin")
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)


def build_rate_card_sheet(wb: Workbook, rate_card: RateCard, entries: list[RateCardEntry]):
    ws = wb.create_sheet("Rate Card")
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20

    _header(ws, 1, 1, "Role")
    _header(ws, 1, 2, "Seniority")
    _header(ws, 1, 3, f"Daily Rate ({rate_card.currency})")

    for i, entry in enumerate(entries, start=2):
        ws.cell(i, 1, entry.role)
        ws.cell(i, 2, entry.seniority)
        ws.cell(i, 3, float(entry.daily_rate))
        ws.cell(i, 3).number_format = '#,##0.00'

    _border_all(ws, 1, 1 + len(entries), 1, 3)
    return ws


def build_wbs_sheet(wb: Workbook, items: list[WBSItem], ws_name: str = "WBS"):
    ws = wb.create_sheet(ws_name)
    for col, width in zip(range(1, 10), [5, 40, 12, 12, 12, 12, 12, 12, 12]):
        ws.column_dimensions[get_column_letter(col)].width = width

    headers = ["#", "Task / Story", "Complexity", "Dev (md)", "QA (md)", "BA (md)", "PM (md)", "Buffer %", "Type"]
    for c, h in enumerate(headers, 1):
        _header(ws, 1, c, h)

    for i, item in enumerate(items, start=2):
        ws.cell(i, 1, i - 1)
        ws.cell(i, 2, item.title)
        ws.cell(i, 3, item.complexity.value if item.complexity else "")
        ws.cell(i, 4, float(item.effort_dev_md or 0))
        ws.cell(i, 5, float(item.effort_qa_md or 0))
        ws.cell(i, 6, float(item.effort_ba_md or 0))
        ws.cell(i, 7, float(item.effort_pm_md or 0))
        ws.cell(i, 8, float(item.risk_buffer_pct or 20))
        ws.cell(i, 9, item.item_type.value)
        if i % 2 == 0:
            for c in range(1, 10):
                ws.cell(i, c).fill = ALT_FILL

    _border_all(ws, 1, 1 + len(items), 1, 9)
    return ws


def build_summary_sheet(
    wb: Workbook,
    items: list[WBSItem],
    rate_card: RateCard,
    entries: list[RateCardEntry],
    wbs_sheet_name: str = "WBS",
    rate_sheet_name: str = "Rate Card",
):
    """
    KL-P-002: cost cells reference WBS effort * Rate Card daily_rate via VLOOKUP formulas.
    Row layout: item# | title | dev_cost | qa_cost | ba_cost | pm_cost | base | buffer | total
    All cost cells = formula, NOT hardcoded.
    """
    ws = wb.create_sheet("Summary")
    for col, width in zip(range(1, 10), [5, 40, 15, 15, 15, 15, 18, 12, 18]):
        ws.column_dimensions[get_column_letter(col)].width = width

    headers = ["#", "Title", "Dev Cost", "QA Cost", "BA Cost", "PM Cost", "Base Cost", "Buffer %", "Total Cost"]
    for c, h in enumerate(headers, 1):
        _header(ws, 1, c, h)

    currency = rate_card.currency
    # Build role->row index map in Rate Card sheet for VLOOKUP
    # Rate Card layout: col A=role, col B=seniority, col C=rate
    # We look up "dev","mid" etc. For simplicity, use first matching mid entry
    role_rows: dict[str, int] = {}
    for i, entry in enumerate(entries, start=2):
        key = f"{entry.role}:{entry.seniority}"
        if key not in role_rows:
            role_rows[key] = i

    dev_row = role_rows.get("dev:mid", 2)
    qa_row = role_rows.get("qa:mid", 3)
    ba_row = role_rows.get("ba:mid", 4)
    pm_row = role_rows.get("pm:mid", 5)

    for i, item in enumerate(items, start=2):
        wbs_row = i  # WBS sheet has same row layout
        ws.cell(i, 1, i - 1)
        ws.cell(i, 2, item.title)

        # KL-P-002: formulas reference cells, not computed values
        dev_formula = f"='{wbs_sheet_name}'!D{wbs_row}*'{rate_sheet_name}'!C{dev_row}"
        qa_formula = f"='{wbs_sheet_name}'!E{wbs_row}*'{rate_sheet_name}'!C{qa_row}"
        ba_formula = f"='{wbs_sheet_name}'!F{wbs_row}*'{rate_sheet_name}'!C{ba_row}"
        pm_formula = f"='{wbs_sheet_name}'!G{wbs_row}*'{rate_sheet_name}'!C{pm_row}"
        base_formula = f"=C{i}+D{i}+E{i}+F{i}"
        buffer_formula = f"='{wbs_sheet_name}'!H{wbs_row}/100"
        total_formula = f"=G{i}*(1+H{i})"

        ws.cell(i, 3, dev_formula)
        ws.cell(i, 4, qa_formula)
        ws.cell(i, 5, ba_formula)
        ws.cell(i, 6, pm_formula)
        ws.cell(i, 7, base_formula)
        ws.cell(i, 8, buffer_formula)
        ws.cell(i, 9, total_formula)

        for c in range(3, 10):
            ws.cell(i, c).number_format = f'#,##0.00 "{currency}"'
        ws.cell(i, 8).number_format = "0.0%"

    # Grand total row
    total_row = 1 + len(items) + 1
    ws.cell(total_row, 2, "GRAND TOTAL").font = BOLD
    if len(items) > 0:
        ws.cell(total_row, 9, f"=SUM(I2:I{1+len(items)})").font = BOLD
        ws.cell(total_row, 9).number_format = f'#,##0.00 "{currency}"'

    _border_all(ws, 1, total_row, 1, 9)
    return ws


def build_test_cases_sheet(wb: Workbook, analysis_results):
    from src.models.analysis_result import ResultType
    ws = wb.create_sheet("Test Cases")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 60

    _header(ws, 1, 1, "#")
    _header(ws, 1, 2, "Title")
    _header(ws, 1, 3, "Severity")
    _header(ws, 1, 4, "Description")

    tc_items = [r for r in analysis_results if r.result_type == ResultType.test_case]
    for i, item in enumerate(tc_items, start=2):
        ws.cell(i, 1, i - 1)
        ws.cell(i, 2, item.title)
        ws.cell(i, 3, item.severity.value if item.severity else "")
        ws.cell(i, 4, item.content)
        ws.cell(i, 4).alignment = Alignment(wrap_text=True)

    _border_all(ws, 1, max(1 + len(tc_items), 2), 1, 4)
    return ws


def export_quotation(
    project_id: uuid.UUID,
    organization_id: uuid.UUID,
    db: Session,
) -> bytes:
    from src.models.analysis_result import AnalysisResult

    rate_card = db.execute(
        select(RateCard).where(
            RateCard.organization_id == organization_id,
            RateCard.is_active == True,
        ).limit(1)
    ).scalar_one_or_none()

    if not rate_card:
        raise ValueError("No active rate card found")

    entries = db.execute(
        select(RateCardEntry).where(RateCardEntry.rate_card_id == rate_card.id)
    ).scalars().all()

    items = db.execute(
        select(WBSItem).where(
            WBSItem.project_id == project_id,
            WBSItem.organization_id == organization_id,
        ).order_by(WBSItem.created_at)
    ).scalars().all()

    analysis_results = db.execute(
        select(AnalysisResult).where(
            AnalysisResult.project_id == project_id,
            AnalysisResult.organization_id == organization_id,
        )
    ).scalars().all()

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_rate_card_sheet(wb, rate_card, entries)
    build_wbs_sheet(wb, items)
    build_summary_sheet(wb, items, rate_card, entries)
    build_test_cases_sheet(wb, analysis_results)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
