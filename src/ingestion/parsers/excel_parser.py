import io
import openpyxl
from .base import BaseParser, ParseResult, ParsedPage


class ExcelParser(BaseParser):
    def parse(self, content: bytes) -> ParseResult:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        pages = []
        for i, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            rows_text = []
            headers: list[str] = []
            header_row = None

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                cells = [str(c) if c is not None else "" for c in row]
                if row_idx == 1:
                    headers = [c for c in cells if c]
                    header_row = " | ".join(headers)
                else:
                    row_text = " | ".join(cells)
                    if any(c.strip() for c in cells):
                        rows_text.append(row_text)

            sheet_text = f"Sheet: {sheet_name}\n"
            if header_row:
                sheet_text += f"Headers: {header_row}\n"
            sheet_text += "\n".join(rows_text)
            pages.append(ParsedPage(page=i, text=sheet_text, headings=[sheet_name]))

        wb.close()
        return ParseResult(pages=pages)
