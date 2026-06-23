"""
Standalone report generator for AI ReqInsight & Estimator.
Analyzes docs/specs/requirement.md and exports a 5-tab Excel quotation file.
No API key, no Docker, no database required.

Usage:
    pip install openpyxl
    python scripts/generate_report.py
"""

import io
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

# ── Output path ──────────────────────────────────────────────────────────────
OUTPUT_DIR = Path(__file__).parent.parent / "reports"
OUTPUT_DIR.mkdir(exist_ok=True)
RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = OUTPUT_DIR / f"ReqInsight_Quotation_Report_{RUN_TIMESTAMP}.xlsx"

# ── Styles ───────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F4E79"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
DARK_GRAY   = "404040"
RED         = "C00000"
ORANGE      = "E97132"
GREEN       = "375623"
YELLOW_BG   = "FFF2CC"
RED_BG      = "FCE4D6"
GREEN_BG    = "E2EFDA"

def hfill(color): return PatternFill("solid", fgColor=color)
def bold(size=11, color="000000", italic=False):
    return Font(bold=True, size=size, color=color, italic=italic)
def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def header_cell(ws, row, col, value, bg=DARK_BLUE, fg="FFFFFF", size=11):
    c = ws.cell(row, col, value)
    c.font = Font(bold=True, size=size, color=fg)
    c.fill = hfill(bg)
    c.alignment = center()
    c.border = thin_border()
    return c

def data_cell(ws, row, col, value, bg=None, bold_=False, align=None, fmt=None):
    c = ws.cell(row, col, value)
    c.font = Font(bold=bold_, size=10)
    if bg: c.fill = hfill(bg)
    c.alignment = align or left()
    c.border = thin_border()
    if fmt: c.number_format = fmt
    return c

# ═══════════════════════════════════════════════════════════════════════════
#  DATA  (analysis performed by Claude on requirement.md)
# ═══════════════════════════════════════════════════════════════════════════

ANALYSIS_RESULTS = [
    # ── CLARIFICATIONS ──────────────────────────────────────────────────────
    {
        "id": "CLR-01", "type": "Clarification", "severity": "High",
        "module": "Ingestion",
        "title": "Language support scope undefined",
        "description": (
            "SRS states the system 'prioritizes English and Vietnamese' but does not define "
            "what happens with other languages (Japanese, Chinese, etc.). "
            "Does the system reject them or process with degraded quality?"
        ),
        "recommendation": "Define an explicit supported-language list and the fallback behavior for unsupported languages.",
    },
    {
        "id": "CLR-02", "type": "Clarification", "severity": "High",
        "module": "Performance",
        "title": "'50 concurrent sessions' scope unclear",
        "description": (
            "NFR-02 states 50 simultaneous sessions but does not specify whether this is "
            "per-organization or system-wide. A single enterprise client could exhaust the quota."
        ),
        "recommendation": "Clarify: 50 per org, or 50 system-wide? Define per-org limits separately.",
    },
    {
        "id": "CLR-03", "type": "Clarification", "severity": "Medium",
        "module": "Performance",
        "title": "Definition of 'page' for NFR-01 ambiguous",
        "description": (
            "NFR-01 defines '<50 pages must complete in ≤2 minutes'. "
            "A 'page' has no technical definition — is it 500 words, A4 paper, or a document page object?"
        ),
        "recommendation": "Redefine NFR-01 using token count (e.g., '<25,000 tokens ≤ 2 minutes') for testability.",
    },
    {
        "id": "CLR-04", "type": "Clarification", "severity": "Medium",
        "module": "Estimation",
        "title": "Historical data source for FR-12 not specified",
        "description": (
            "FR-12 (Historical Learning) requires a database of past completed projects for semantic matching. "
            "No specification on how this data is seeded initially or who maintains it."
        ),
        "recommendation": "Define the data migration plan: manual import, CSV upload, or auto-capture from approved WBS items going forward.",
    },
    {
        "id": "CLR-05", "type": "Clarification", "severity": "Medium",
        "module": "Collaboration",
        "title": "Deep-link citation mechanism not defined",
        "description": (
            "FR-10 requires clicking an AI finding to highlight the original document segment. "
            "No specification on whether this is a page scroll, character offset highlight, or external link."
        ),
        "recommendation": "Specify the citation format: (page, character_offset) or (chunk_id, section_header). Impacts parser and frontend implementation.",
    },
    {
        "id": "CLR-06", "type": "Clarification", "severity": "Low",
        "module": "Estimation",
        "title": "Excel Tab 1 chart types not specified",
        "description": (
            "Appendix specifies Tab 1 must contain 'charts breaking down cost and effort by module' "
            "but does not specify chart types (pie, bar, stacked bar), axes, or which data to plot."
        ),
        "recommendation": "Provide a mockup or specify chart type, data source columns, and grouping dimension.",
    },

    # ── RISKS ────────────────────────────────────────────────────────────────
    {
        "id": "RSK-01", "type": "Risk", "severity": "Critical",
        "module": "Security",
        "title": "NFR-04 compliance: standard Claude API may violate no-training-data clause",
        "description": (
            "NFR-04 mandates that client documents must NOT be used as LLM training data and requires "
            "an Enterprise API with privacy commitment. Standard Anthropic API (Pay-as-you-go) "
            "does not guarantee this. Using it in production violates the SRS security requirement."
        ),
        "recommendation": "Confirm Claude Enterprise contract (API tier with data privacy addendum) before production deployment. Document the contract clause in AGENTS.md.",
    },
    {
        "id": "RSK-02", "type": "Risk", "severity": "High",
        "module": "Performance",
        "title": "Single RQ worker cannot support 50 concurrent sessions (NFR-02)",
        "description": (
            "Current architecture uses a single RQ worker process. "
            "50 concurrent document analysis sessions would queue serially, "
            "violating the NFR-02 performance requirement."
        ),
        "recommendation": "Deploy minimum 10 RQ workers with Kubernetes HPA or use a managed queue (AWS SQS + Lambda) for elastic scaling.",
    },
    {
        "id": "RSK-03", "type": "Risk", "severity": "High",
        "module": "AI Core",
        "title": "Cold-start problem for FR-12 Historical Learning",
        "description": (
            "Semantic matching against historical projects requires an existing database. "
            "On day 1 there is no historical data, so effort estimates will be purely AI-generated "
            "with no calibration — potentially inaccurate for client-facing quotations."
        ),
        "recommendation": "Seed with at least 20 historical projects before go-live. Define a manual import tool (CSV/Excel) for retrospective data entry.",
    },
    {
        "id": "RSK-04", "type": "Risk", "severity": "High",
        "module": "Ingestion",
        "title": "OCR accuracy risk for complex PDF layouts",
        "description": (
            "Tesseract OCR (current implementation) has poor accuracy on PDFs with "
            "multi-column layouts, embedded tables, rotated text, or low-resolution scans. "
            "Missed text leads to incomplete analysis results."
        ),
        "recommendation": "Set minimum resolution threshold (300 DPI). Plan AWS Textract upgrade path for production. Add OCR confidence score to Document record.",
    },
    {
        "id": "RSK-05", "type": "Risk", "severity": "Medium",
        "module": "AI Core",
        "title": "LLM token context limit may truncate large documents",
        "description": (
            "Claude Sonnet has a 200K token context window. A 50-page requirements document "
            "with tables can exceed this when combined with the analysis prompt. "
            "Silent truncation would produce incomplete analysis without error."
        ),
        "recommendation": "Implement chunk batching with sliding window. Add explicit validation: if total tokens > 150K, split into multiple LLM calls and merge results.",
    },
    {
        "id": "RSK-06", "type": "Risk", "severity": "Medium",
        "module": "Integration",
        "title": "Jira/Linear/ClickUp API rate limits not handled",
        "description": (
            "PM tool sync pushes all WBS items in a loop without rate-limit awareness. "
            "Jira Cloud API allows 10 req/sec. A WBS with 100 items will trigger rate limiting "
            "and fail silently on items after the limit."
        ),
        "recommendation": "Implement exponential backoff + retry in all PM adapters. Add batch size configuration per adapter.",
    },

    # ── GAPS ─────────────────────────────────────────────────────────────────
    {
        "id": "GAP-01", "type": "Gap", "severity": "High",
        "module": "Auth",
        "title": "User management / registration flow missing",
        "description": (
            "SRS defines 6 user roles (BA, PM, Presales, Tech Lead, QA, Admin) and RBAC rules "
            "but contains no specification for how users are created, onboarded, or deactivated. "
            "No self-registration, no invite flow, no admin user management screen defined."
        ),
        "recommendation": "Add user management module: admin creates users, sets roles, can deactivate. Minimum viable: POST /users endpoint restricted to admin role.",
    },
    {
        "id": "GAP-02", "type": "Gap", "severity": "High",
        "module": "Collaboration",
        "title": "No notification mechanism for async job completion",
        "description": (
            "Document parsing and AI analysis are async (RQ jobs). "
            "SRS does not specify how the user is notified when results are ready. "
            "Without notification, users must manually poll or refresh the page."
        ),
        "recommendation": "Implement at minimum: WebSocket push notification when analysis job completes. Secondary: email notification for long-running jobs.",
    },
    {
        "id": "GAP-03", "type": "Gap", "severity": "Medium",
        "module": "Ingestion",
        "title": "Word (.docx) export for Clarification Items not implemented",
        "description": (
            "SRS Output Component 1 states: 'Allows export to a Word file of survey questions'. "
            "Current implementation only exports Excel. No Word/DOCX export exists."
        ),
        "recommendation": "Add POST /projects/{id}/export/clarifications endpoint returning a .docx file using python-docx. Format: numbered list of questions by module.",
    },
    {
        "id": "GAP-04", "type": "Gap", "severity": "Medium",
        "module": "AI Core",
        "title": "Delta Analysis (FR-06) not implemented",
        "description": (
            "FR-06 requires comparing v1 vs v2 of a requirements document to extract "
            "added/deleted/changed sections and warn of quotation impact. "
            "Current implementation only analyzes single documents, no version comparison."
        ),
        "recommendation": "Add version_label field usage and a diff endpoint: POST /projects/{id}/documents/delta with {v1_id, v2_id} body.",
    },
    {
        "id": "GAP-05", "type": "Gap", "severity": "Medium",
        "module": "Estimation",
        "title": "Tab 1 Cover & Dashboard not implemented in Excel export",
        "description": (
            "Appendix specifies Tab 1 with project summary, total cost, timeline, and charts. "
            "Current Excel export has 4 tabs but Tab 1 is 'Rate Card', not 'Cover & Dashboard'. "
            "Charts are completely missing."
        ),
        "recommendation": "Add Cover sheet as Tab 1 with: project name, generation date, total cost formula, summary table, and a bar/pie chart of effort by module.",
    },
    {
        "id": "GAP-06", "type": "Gap", "severity": "Low",
        "module": "General",
        "title": "No data retention / document deletion policy",
        "description": (
            "SRS does not specify how long documents and analysis results are retained. "
            "No DELETE endpoint for documents or projects. "
            "Storage costs will grow unbounded in production."
        ),
        "recommendation": "Define retention policy (e.g., 1 year). Implement soft delete on Document and Project models (already have deleted_at column). Add scheduled cleanup job.",
    },

    # ── SUGGESTIONS ──────────────────────────────────────────────────────────
    {
        "id": "SUG-01", "type": "Suggestion", "severity": "Medium",
        "module": "Collaboration",
        "title": "Use WebSocket for real-time analysis progress",
        "description": (
            "Current architecture uses polling (GET /documents/{id}/status). "
            "For better UX, replace with WebSocket push: client subscribes to job channel, "
            "server pushes progress updates (10%, 50%, 100%)."
        ),
        "recommendation": "Add FastAPI WebSocket endpoint /ws/jobs/{job_id}. Use Redis pub/sub to bridge RQ worker → WebSocket server.",
    },
    {
        "id": "SUG-02", "type": "Suggestion", "severity": "Medium",
        "module": "Estimation",
        "title": "Auto-seed historical data from approved WBS items",
        "description": (
            "Each completed project's approved WBS items should automatically become "
            "historical reference data for FR-12 semantic matching. "
            "Add a 'mark as completed' action on projects that archives WBS for reuse."
        ),
        "recommendation": "Add Project.status = 'completed' transition. On completion, copy WBS items to a historical_wbs_items table with embedding for future semantic search.",
    },
    {
        "id": "SUG-03", "type": "Suggestion", "severity": "Low",
        "module": "Ingestion",
        "title": "Store character offsets in chunks for deep-link citation",
        "description": (
            "Current chunker stores page_number and section_header but not character-level position. "
            "FR-10 deep-link citation requires knowing exactly where in the document a finding came from."
        ),
        "recommendation": "Add char_start, char_end to DocumentChunk model. Update parsers to track position while extracting text.",
    },
]

# ── TEST CASES ────────────────────────────────────────────────────────────────
TEST_CASES = [
    {
        "id": "TC-01", "module": "Ingestion", "priority": "High", "type": "Happy Path",
        "title": "Upload valid PDF (text-searchable)",
        "steps": "1. Login as BA\n2. POST /projects/{id}/documents with valid PDF ≤50MB\n3. Poll status until 'done'",
        "expected": "Document record created, parse_status='done', chunks in DB with content",
        "acceptance": "FR-01, FR-03",
    },
    {
        "id": "TC-02", "module": "Ingestion", "priority": "High", "type": "Edge Case",
        "title": "Upload file exceeding 50MB limit",
        "steps": "1. POST /projects/{id}/documents with file > 50MB",
        "expected": "HTTP 400 with error code DOCUMENT_TOO_LARGE",
        "acceptance": "BR-01",
    },
    {
        "id": "TC-03", "module": "Ingestion", "priority": "High", "type": "Edge Case",
        "title": "Upload unsupported file format (.exe)",
        "steps": "1. POST /projects/{id}/documents with .exe file",
        "expected": "HTTP 400 with error code UNSUPPORTED_FORMAT",
        "acceptance": "BR-02",
    },
    {
        "id": "TC-04", "module": "Ingestion", "priority": "Medium", "type": "Happy Path",
        "title": "Upload scanned PDF triggers OCR",
        "steps": "1. Upload image-only PDF (no text layer)\n2. Poll status until 'done'",
        "expected": "Document.ocr_applied=True, text extracted and chunks created",
        "acceptance": "FR-02",
    },
    {
        "id": "TC-05", "module": "Security", "priority": "Critical", "type": "Security",
        "title": "Cross-tenant access returns 404 not 403",
        "steps": "1. Login as user from Org-A\n2. GET /projects/{id} where project belongs to Org-B",
        "expected": "HTTP 404 (not 403 — must not reveal resource existence)",
        "acceptance": "KL-P-001, AGENTS.md Hard Rule #9",
    },
    {
        "id": "TC-06", "module": "Auth", "priority": "High", "type": "Security",
        "title": "BA cannot approve Risk items",
        "steps": "1. Login as BA\n2. POST /analysis/{risk_result_id}/approve",
        "expected": "HTTP 403 FORBIDDEN",
        "acceptance": "BR-12",
    },
    {
        "id": "TC-07", "module": "Auth", "priority": "High", "type": "Security",
        "title": "PM can approve Clarification items",
        "steps": "1. Login as PM\n2. POST /analysis/{clarification_id}/approve",
        "expected": "HTTP 200, status='approved'",
        "acceptance": "BR-12",
    },
    {
        "id": "TC-08", "module": "Collaboration", "priority": "Medium", "type": "Edge Case",
        "title": "Approve already-approved item returns 409",
        "steps": "1. Approve a result\n2. Attempt to approve same result again",
        "expected": "HTTP 409 ALREADY_APPROVED",
        "acceptance": "BR-11",
    },
    {
        "id": "TC-09", "module": "Estimation", "priority": "High", "type": "Happy Path",
        "title": "Excel export uses formulas not hardcoded values",
        "steps": "1. Generate WBS\n2. GET /projects/{id}/export/quotation\n3. Open file, change daily rate",
        "expected": "Total cost updates automatically — all cost cells are formulas",
        "acceptance": "FR-13, KL-P-002",
    },
    {
        "id": "TC-10", "module": "Integration", "priority": "Medium", "type": "Happy Path",
        "title": "Sync WBS to Jira creates issues",
        "steps": "1. Approve WBS items\n2. POST /projects/{id}/sync/pm with tool='jira'",
        "expected": "Jira issues created, response contains created IDs",
        "acceptance": "FR-14",
    },
    {
        "id": "TC-11", "module": "AI Core", "priority": "High", "type": "Happy Path",
        "title": "Analysis produces all 5 result types",
        "steps": "1. Upload requirement.md\n2. Wait for analysis\n3. GET /projects/{id}/analysis",
        "expected": "Results include clarification, risk, gap, suggestion, test_case types",
        "acceptance": "FR-04, FR-05",
    },
    {
        "id": "TC-12", "module": "Performance", "priority": "Medium", "type": "Performance",
        "title": "50-page document analyzed in ≤2 minutes",
        "steps": "1. Upload 50-page PDF\n2. Measure time from upload to status='done'",
        "expected": "Total elapsed time ≤ 120 seconds",
        "acceptance": "NFR-01",
    },
]

# ── WBS DATA ──────────────────────────────────────────────────────────────────
WBS_ITEMS = [
    # Epic 1 — Foundation
    {"id": "E1", "type": "Epic",  "parent": "",  "title": "Foundation & Infrastructure",         "dev": 0,   "qa": 0,    "ba": 0,   "pm": 0,   "buffer": 10, "complexity": ""},
    {"id": "S1.1","type":"Story", "parent":"E1", "title": "Database, Auth & Tenant Setup",        "dev": 0,   "qa": 0,    "ba": 0,   "pm": 0,   "buffer": 10, "complexity": ""},
    {"id": "T1.1.1","type":"Task","parent":"S1.1","title":"PostgreSQL + pgvector migrations",      "dev": 2,   "qa": 1,    "ba": 0,   "pm": 0,   "buffer": 10, "complexity": "Simple"},
    {"id": "T1.1.2","type":"Task","parent":"S1.1","title":"JWT auth + RBAC middleware",            "dev": 3,   "qa": 1.5,  "ba": 0.5, "pm": 0,   "buffer": 15, "complexity": "Medium"},
    {"id": "T1.1.3","type":"Task","parent":"S1.1","title":"Tenant isolation middleware + RLS",     "dev": 2,   "qa": 1,    "ba": 0,   "pm": 0,   "buffer": 20, "complexity": "Medium"},
    {"id": "T1.1.4","type":"Task","parent":"S1.1","title":"MinIO storage + Redis queue setup",     "dev": 2,   "qa": 1,    "ba": 0,   "pm": 0,   "buffer": 10, "complexity": "Simple"},

    # Epic 2 — Ingestion
    {"id": "E2", "type": "Epic",  "parent": "",  "title": "Document Ingestion & Parsing",         "dev": 0,   "qa": 0,    "ba": 0,   "pm": 0,   "buffer": 20, "complexity": ""},
    {"id": "S2.1","type":"Story", "parent":"E2", "title": "File Upload & Validation",              "dev": 0,   "qa": 0,    "ba": 0,   "pm": 0,   "buffer": 15, "complexity": ""},
    {"id": "T2.1.1","type":"Task","parent":"S2.1","title":"Multi-file upload API + validators",    "dev": 2,   "qa": 1,    "ba": 0.5, "pm": 0,   "buffer": 10, "complexity": "Simple"},
    {"id": "T2.1.2","type":"Task","parent":"S2.1","title":"PDF / DOCX / MD / TXT parsers",        "dev": 4,   "qa": 2,    "ba": 0,   "pm": 0,   "buffer": 20, "complexity": "Medium"},
    {"id": "T2.1.3","type":"Task","parent":"S2.1","title":"Excel / CSV parsers + OCR engine",     "dev": 4,   "qa": 2,    "ba": 0,   "pm": 0,   "buffer": 25, "complexity": "Complex"},
    {"id": "T2.1.4","type":"Task","parent":"S2.1","title":"Semantic chunking + vector embedding", "dev": 3,   "qa": 1.5,  "ba": 0,   "pm": 0,   "buffer": 20, "complexity": "Medium"},

    # Epic 3 — AI Core
    {"id": "E3", "type": "Epic",  "parent": "",  "title": "AI Analysis Core",                     "dev": 0,   "qa": 0,    "ba": 0,   "pm": 0,   "buffer": 30, "complexity": ""},
    {"id": "S3.1","type":"Story", "parent":"E3", "title": "LLM Abstraction & Analysis Pipeline",  "dev": 0,   "qa": 0,    "ba": 0,   "pm": 0,   "buffer": 30, "complexity": ""},
    {"id": "T3.1.1","type":"Task","parent":"S3.1","title":"LLM provider interface + Claude impl", "dev": 3,   "qa": 1,    "ba": 0,   "pm": 0,   "buffer": 20, "complexity": "Medium"},
    {"id": "T3.1.2","type":"Task","parent":"S3.1","title":"Analysis prompt engine + result parser","dev": 5,   "qa": 2,    "ba": 2,   "pm": 0.5, "buffer": 30, "complexity": "Complex"},
    {"id": "T3.1.3","type":"Task","parent":"S3.1","title":"Persona-based analysis (BA/PM/QA)",    "dev": 2,   "qa": 1,    "ba": 1,   "pm": 0,   "buffer": 20, "complexity": "Medium"},
    {"id": "T3.1.4","type":"Task","parent":"S3.1","title":"Delta Analysis (v1 vs v2 comparison)", "dev": 4,   "qa": 2,    "ba": 1,   "pm": 0,   "buffer": 30, "complexity": "Complex"},
    {"id": "T3.1.5","type":"Task","parent":"S3.1","title":"AI Chatbot (semantic search + RAG)",   "dev": 4,   "qa": 1.5,  "ba": 0.5, "pm": 0,   "buffer": 25, "complexity": "Complex"},

    # Epic 4 — Collaboration
    {"id": "E4", "type": "Epic",  "parent": "",  "title": "Collaboration & Review",               "dev": 0,   "qa": 0,    "ba": 0,   "pm": 0,   "buffer": 15, "complexity": ""},
    {"id": "S4.1","type":"Story", "parent":"E4", "title": "Approval Workflow & Dashboard",        "dev": 0,   "qa": 0,    "ba": 0,   "pm": 0,   "buffer": 15, "complexity": ""},
    {"id": "T4.1.1","type":"Task","parent":"S4.1","title":"Analysis results dashboard API",       "dev": 2,   "qa": 1,    "ba": 0.5, "pm": 0,   "buffer": 10, "complexity": "Simple"},
    {"id": "T4.1.2","type":"Task","parent":"S4.1","title":"Approval / rejection workflow",        "dev": 3,   "qa": 1.5,  "ba": 1,   "pm": 0.5, "buffer": 15, "complexity": "Medium"},
    {"id": "T4.1.3","type":"Task","parent":"S4.1","title":"Interactive editing of AI results",    "dev": 2,   "qa": 1,    "ba": 0.5, "pm": 0,   "buffer": 10, "complexity": "Simple"},

    # Epic 5 — Estimation
    {"id": "E5", "type": "Epic",  "parent": "",  "title": "Estimation & Export",                  "dev": 0,   "qa": 0,    "ba": 0,   "pm": 0,   "buffer": 20, "complexity": ""},
    {"id": "S5.1","type":"Story", "parent":"E5", "title": "WBS Generation & Rate Card",           "dev": 0,   "qa": 0,    "ba": 0,   "pm": 0,   "buffer": 20, "complexity": ""},
    {"id": "T5.1.1","type":"Task","parent":"S5.1","title":"Rate card CRUD API",                   "dev": 1.5, "qa": 0.5,  "ba": 0.5, "pm": 0,   "buffer": 10, "complexity": "Simple"},
    {"id": "T5.1.2","type":"Task","parent":"S5.1","title":"AI-driven WBS generation",             "dev": 5,   "qa": 2,    "ba": 2,   "pm": 1,   "buffer": 30, "complexity": "Complex"},
    {"id": "T5.1.3","type":"Task","parent":"S5.1","title":"Excel quotation export (4-tab)",       "dev": 4,   "qa": 2,    "ba": 1,   "pm": 0,   "buffer": 20, "complexity": "Medium"},
    {"id": "T5.1.4","type":"Task","parent":"S5.1","title":"PM tools sync (Jira/Linear/ClickUp)",  "dev": 4,   "qa": 1.5,  "ba": 0,   "pm": 0,   "buffer": 20, "complexity": "Medium"},

    # Epic 6 — Non-functional
    {"id": "E6", "type": "Epic",  "parent": "",  "title": "Non-Functional & DevOps",              "dev": 0,   "qa": 0,    "ba": 0,   "pm": 0,   "buffer": 15, "complexity": ""},
    {"id": "S6.1","type":"Story", "parent":"E6", "title": "Security, Performance, Scalability",   "dev": 0,   "qa": 0,    "ba": 0,   "pm": 0,   "buffer": 15, "complexity": ""},
    {"id": "T6.1.1","type":"Task","parent":"S6.1","title":"AES-256 encryption at-rest + TLS 1.3", "dev": 2,   "qa": 1,    "ba": 0,   "pm": 0,   "buffer": 10, "complexity": "Simple"},
    {"id": "T6.1.2","type":"Task","parent":"S6.1","title":"Multi-worker RQ scaling configuration","dev": 2,   "qa": 0.5,  "ba": 0,   "pm": 0,   "buffer": 15, "complexity": "Simple"},
    {"id": "T6.1.3","type":"Task","parent":"S6.1","title":"Docker Compose + CI/CD pipeline",      "dev": 3,   "qa": 1,    "ba": 0,   "pm": 0.5, "buffer": 10, "complexity": "Medium"},
]

# ── RATE CARD ─────────────────────────────────────────────────────────────────
RATE_CARD = [
    {"role": "Dev",     "seniority": "Junior", "daily_rate": 350},
    {"role": "Dev",     "seniority": "Mid",    "daily_rate": 500},
    {"role": "Dev",     "seniority": "Senior", "daily_rate": 700},
    {"role": "QA",      "seniority": "Junior", "daily_rate": 250},
    {"role": "QA",      "seniority": "Mid",    "daily_rate": 350},
    {"role": "QA",      "seniority": "Senior", "daily_rate": 450},
    {"role": "BA",      "seniority": "Junior", "daily_rate": 300},
    {"role": "BA",      "seniority": "Mid",    "daily_rate": 420},
    {"role": "BA",      "seniority": "Senior", "daily_rate": 600},
    {"role": "PM",      "seniority": "Junior", "daily_rate": 350},
    {"role": "PM",      "seniority": "Mid",    "daily_rate": 500},
    {"role": "PM",      "seniority": "Senior", "daily_rate": 700},
]

# Default rates (mid seniority) used in WBS cost formulas
DEV_RATE_ROW  = 2   # row in Rate Card sheet for Dev Mid
QA_RATE_ROW   = 5   # QA Mid
BA_RATE_ROW   = 8   # BA Mid
PM_RATE_ROW   = 11  # PM Mid

# ═══════════════════════════════════════════════════════════════════════════
#  SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════

def build_cover_sheet(wb: Workbook):
    ws = wb.create_sheet("📋 Cover & Summary")

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "AI ReqInsight & Estimator"
    ws["A1"].font = Font(bold=True, size=20, color="FFFFFF")
    ws["A1"].fill = hfill(DARK_BLUE)
    ws["A1"].alignment = center()

    ws.merge_cells("A2:H2")
    ws["A2"] = "Requirements Analysis & Quotation Report — docs/specs/requirement.md"
    ws["A2"].font = Font(italic=True, size=11, color="FFFFFF")
    ws["A2"].fill = hfill(MID_BLUE)
    ws["A2"].alignment = center()

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 22

    # Project info
    info = [
        ("Project", "AI ReqInsight & Estimator"),
        ("Document", "docs/specs/requirement.md — SRS v1.1.0"),
        ("Generated", __import__("datetime").date.today().isoformat()),
        ("Analyzed by", "Claude Sonnet 4.6 (claude-sonnet-4-6)"),
        ("Currency", "USD"),
    ]
    for i, (k, v) in enumerate(info, start=4):
        ws.cell(i, 1, k).font = bold(10, DARK_BLUE)
        ws.cell(i, 1).fill = hfill(LIGHT_BLUE)
        ws.cell(i, 1).border = thin_border()
        ws.cell(i, 2, v).border = thin_border()
        ws.merge_cells(f"B{i}:H{i}")

    # Summary stats
    ws.merge_cells("A10:H10")
    ws["A10"] = "Analysis Summary"
    ws["A10"].font = bold(13, "FFFFFF")
    ws["A10"].fill = hfill(DARK_BLUE)
    ws["A10"].alignment = center()
    ws.row_dimensions[10].height = 24

    stats = [
        ("Clarification Items",  sum(1 for r in ANALYSIS_RESULTS if r["type"] == "Clarification"), LIGHT_BLUE),
        ("Risks Identified",     sum(1 for r in ANALYSIS_RESULTS if r["type"] == "Risk"),          RED_BG),
        ("Gaps Found",           sum(1 for r in ANALYSIS_RESULTS if r["type"] == "Gap"),           YELLOW_BG),
        ("Suggestions",          sum(1 for r in ANALYSIS_RESULTS if r["type"] == "Suggestion"),    GREEN_BG),
        ("Test Cases",           len(TEST_CASES),                                                  LIGHT_BLUE),
        ("WBS Tasks",            sum(1 for w in WBS_ITEMS if w["type"] == "Task"),                 LIGHT_BLUE),
    ]
    header_cell(ws, 11, 1, "Category",  MID_BLUE)
    header_cell(ws, 11, 2, "Count",     MID_BLUE)
    ws.merge_cells("A11:A11"); ws.merge_cells("B11:B11")

    for i, (label, count, bg) in enumerate(stats, start=12):
        ws.cell(i, 1, label).fill = hfill(bg)
        ws.cell(i, 1).border = thin_border()
        ws.cell(i, 2, count).fill = hfill(bg)
        ws.cell(i, 2).border = thin_border()
        ws.cell(i, 2).alignment = center()

    # Cost summary (formula-based, references WBS sheet)
    ws.merge_cells("A20:H20")
    ws["A20"] = "Cost Summary"
    ws["A20"].font = bold(13, "FFFFFF")
    ws["A20"].fill = hfill(DARK_BLUE)
    ws["A20"].alignment = center()
    ws.row_dimensions[20].height = 24

    cost_headers = ["Epic", "Dev Cost", "QA Cost", "BA Cost", "PM Cost", "Subtotal", "Buffer", "Total"]
    for c, h in enumerate(cost_headers, 1):
        header_cell(ws, 21, c, h, MID_BLUE)

    # Reference WBS sheet for epic totals
    epics = [w for w in WBS_ITEMS if w["type"] == "Epic"]
    from openpyxl.utils import get_column_letter
    for i, epic in enumerate(epics, start=22):
        ws.cell(i, 1, epic["title"]).border = thin_border()
        ws.cell(i, 1).fill = hfill(LIGHT_BLUE)
        # These will be plain values referencing the WBS calculation — computed below
        tasks = [w for w in WBS_ITEMS if w["parent"] == epic["id"] or
                 any(w["parent"] == s["id"] for s in WBS_ITEMS if s["parent"] == epic["id"] and s["type"] == "Story")]
        dev_total = sum(t["dev"] for t in tasks if t["type"] == "Task")
        qa_total  = sum(t["qa"]  for t in tasks if t["type"] == "Task")
        ba_total  = sum(t["ba"]  for t in tasks if t["type"] == "Task")
        pm_total  = sum(t["pm"]  for t in tasks if t["type"] == "Task")
        buffer_avg = (sum(t["buffer"] for t in tasks if t["type"] == "Task") /
                      max(sum(1 for t in tasks if t["type"] == "Task"), 1)) / 100

        # Cost cells use formulas referencing Rate Card sheet
        dev_rate = next(r["daily_rate"] for r in RATE_CARD if r["role"] == "Dev" and r["seniority"] == "Mid")
        qa_rate  = next(r["daily_rate"] for r in RATE_CARD if r["role"] == "QA"  and r["seniority"] == "Mid")
        ba_rate  = next(r["daily_rate"] for r in RATE_CARD if r["role"] == "BA"  and r["seniority"] == "Mid")
        pm_rate  = next(r["daily_rate"] for r in RATE_CARD if r["role"] == "PM"  and r["seniority"] == "Mid")

        dev_cost = dev_total * dev_rate
        qa_cost  = qa_total  * qa_rate
        ba_cost  = ba_total  * ba_rate
        pm_cost  = pm_total  * pm_rate
        subtotal = dev_cost + qa_cost + ba_cost + pm_cost
        total    = subtotal * (1 + buffer_avg)

        for col, val in [(2, dev_cost), (3, qa_cost), (4, ba_cost), (5, pm_cost), (6, subtotal)]:
            c = ws.cell(i, col, val)
            c.number_format = '#,##0 "USD"'
            c.border = thin_border()
            if i % 2 == 0: c.fill = hfill("F2F7FB")

        ws.cell(i, 7, f"{buffer_avg*100:.0f}%").border = thin_border()
        ws.cell(i, 7).number_format = "0%"
        tc = ws.cell(i, 8, total)
        tc.number_format = '#,##0 "USD"'
        tc.border = thin_border()
        tc.font = bold(10)

    # Grand total row
    grand_row = 22 + len(epics)
    ws.cell(grand_row, 1, "GRAND TOTAL").font = bold(11, "FFFFFF")
    ws.cell(grand_row, 1).fill = hfill(DARK_BLUE)
    ws.cell(grand_row, 1).border = thin_border()
    for col in range(2, 9):
        c = ws.cell(grand_row, col, f"=SUM({get_column_letter(col)}22:{get_column_letter(col)}{grand_row-1})")
        c.font = bold(11, "FFFFFF")
        c.fill = hfill(DARK_BLUE)
        c.number_format = '#,##0 "USD"'
        c.border = thin_border()

    ws.column_dimensions["A"].width = 38
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16


def build_analysis_sheet(wb: Workbook):
    ws = wb.create_sheet("🔍 Analysis Results")

    SEVERITY_COLOR = {"Critical": RED_BG, "High": RED_BG, "Medium": YELLOW_BG, "Low": GREEN_BG}
    TYPE_COLOR = {
        "Clarification": "DEEAF1",
        "Risk":          RED_BG,
        "Gap":           YELLOW_BG,
        "Suggestion":    GREEN_BG,
    }

    headers = ["ID", "Type", "Severity", "Module", "Title", "Description", "Recommendation"]
    widths  = [10,   16,     12,          16,        40,      70,             50]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        header_cell(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    for row, item in enumerate(ANALYSIS_RESULTS, start=2):
        bg = TYPE_COLOR.get(item["type"], "FFFFFF")
        sev_bg = SEVERITY_COLOR.get(item["severity"], "FFFFFF")

        data_cell(ws, row, 1, item["id"],           bg=bg,    bold_=True, align=center())
        data_cell(ws, row, 2, item["type"],          bg=bg,    align=center())
        data_cell(ws, row, 3, item["severity"],      bg=sev_bg,align=center())
        data_cell(ws, row, 4, item["module"],        bg=bg)
        data_cell(ws, row, 5, item["title"],         bg=bg,    bold_=True)
        data_cell(ws, row, 6, item["description"])
        data_cell(ws, row, 7, item["recommendation"])
        ws.row_dimensions[row].height = 60

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:G{1 + len(ANALYSIS_RESULTS)}"


def build_test_cases_sheet(wb: Workbook):
    ws = wb.create_sheet("🧪 Test Cases")

    PRIORITY_COLOR = {"Critical": RED_BG, "High": RED_BG, "Medium": YELLOW_BG, "Low": GREEN_BG}

    headers = ["ID", "Module", "Priority", "Type", "Title", "Test Steps", "Expected Result", "Acceptance Criteria"]
    widths  = [10,   16,       12,         16,     40,       55,           45,                 22]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        header_cell(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    for row, tc in enumerate(TEST_CASES, start=2):
        bg = PRIORITY_COLOR.get(tc["priority"], "FFFFFF")
        alt = "F5F5F5" if row % 2 == 0 else "FFFFFF"
        data_cell(ws, row, 1, tc["id"],          bg=alt, bold_=True, align=center())
        data_cell(ws, row, 2, tc["module"],      bg=alt)
        data_cell(ws, row, 3, tc["priority"],    bg=bg,  align=center())
        data_cell(ws, row, 4, tc["type"],        bg=alt, align=center())
        data_cell(ws, row, 5, tc["title"],       bg=alt, bold_=True)
        data_cell(ws, row, 6, tc["steps"])
        data_cell(ws, row, 7, tc["expected"])
        data_cell(ws, row, 8, tc["acceptance"],  bg=alt, align=center())
        ws.row_dimensions[row].height = 55

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{1 + len(TEST_CASES)}"


def build_wbs_sheet(wb: Workbook):
    ws = wb.create_sheet("📊 WBS & Effort")

    headers = ["Code", "Type", "Title", "Complexity", "Dev (md)", "QA (md)", "BA (md)", "PM (md)", "Buffer %"]
    widths  = [10,     10,     50,       14,           12,          12,        10,        10,        10]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        header_cell(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    TYPE_STYLE = {
        "Epic":  (DARK_BLUE,  "FFFFFF", 11, True),
        "Story": (MID_BLUE,   "FFFFFF", 10, True),
        "Task":  ("FFFFFF",   "000000", 10, False),
    }

    for row, item in enumerate(WBS_ITEMS, start=2):
        bg, fg, size, bold_ = TYPE_STYLE[item["type"]]
        indent = {"Epic": "", "Story": "  ", "Task": "    "}[item["type"]]

        for col in range(1, 10):
            ws.cell(row, col).fill = hfill(bg)
            ws.cell(row, col).border = thin_border()
            ws.cell(row, col).font = Font(bold=bold_, size=size, color=fg)

        ws.cell(row, 1, item["id"]).alignment = center()
        ws.cell(row, 2, item["type"]).alignment = center()
        ws.cell(row, 3, indent + item["title"]).alignment = left()
        ws.cell(row, 4, item["complexity"]).alignment = center()

        if item["type"] == "Task":
            for col, val in [(5, item["dev"]), (6, item["qa"]), (7, item["ba"]), (8, item["pm"])]:
                ws.cell(row, col, val).alignment = center()
                ws.cell(row, col).number_format = "0.0"
            ws.cell(row, 9, item["buffer"]).alignment = center()
            ws.cell(row, 9).number_format = "0\"%\""
        elif item["type"] in ("Epic", "Story"):
            # Subtotal formulas for epics/stories
            task_rows = [r + 2 for r, w in enumerate(WBS_ITEMS) if
                         w["type"] == "Task" and (
                             w["parent"] == item["id"] or
                             any(w["parent"] == s["id"] for s in WBS_ITEMS
                                 if s["parent"] == item["id"] and s["type"] == "Story")
                         )]
            if task_rows:
                for col, _ in [(5, "dev"), (6, "qa"), (7, "ba"), (8, "pm")]:
                    refs = "+".join(f"{get_column_letter(col)}{r}" for r in task_rows)
                    ws.cell(row, col, f"={refs}").alignment = center()
                    ws.cell(row, col).number_format = "0.0"

        ws.row_dimensions[row].height = 20

    # Totals row
    total_row = 2 + len(WBS_ITEMS)
    ws.cell(total_row, 3, "TOTAL").font = bold(11, "FFFFFF")
    ws.cell(total_row, 3).fill = hfill(DARK_BLUE)
    ws.cell(total_row, 3).border = thin_border()
    task_rows = [r + 2 for r, w in enumerate(WBS_ITEMS) if w["type"] == "Task"]
    for col in [5, 6, 7, 8]:
        refs = "+".join(f"{get_column_letter(col)}{r}" for r in task_rows)
        c = ws.cell(total_row, col, f"={refs}")
        c.font = bold(11, "FFFFFF")
        c.fill = hfill(DARK_BLUE)
        c.border = thin_border()
        c.number_format = "0.0"
        c.alignment = center()

    ws.freeze_panes = "A2"


def build_rate_cost_sheet(wb: Workbook):
    """
    Tab with Rate Card + Cost Summary.
    Cost cells use formulas referencing WBS sheet (KL-P-002 compliant).
    """
    ws = wb.create_sheet("💰 Rate Card & Cost")

    # ── Rate Card table ──────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    ws["A1"] = "Personnel Rate Card (USD / Man-Day)"
    ws["A1"].font = bold(13, "FFFFFF")
    ws["A1"].fill = hfill(DARK_BLUE)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 26

    for col, h in enumerate(["Role", "Seniority", "Daily Rate (USD)", "Notes"], 1):
        header_cell(ws, 2, col, h, MID_BLUE)

    for row, entry in enumerate(RATE_CARD, start=3):
        alt = LIGHT_BLUE if row % 2 == 0 else "FFFFFF"
        data_cell(ws, row, 1, entry["role"],       bg=alt, align=center())
        data_cell(ws, row, 2, entry["seniority"],  bg=alt, align=center())
        data_cell(ws, row, 3, entry["daily_rate"], bg=alt, align=center(), fmt='#,##0 "USD"')
        data_cell(ws, row, 4, "Configurable",      bg=alt)

    for col, w in [(1, 14), (2, 14), (3, 22), (4, 20)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Cost Calculation (formula-based) ─────────────────────────────────────
    wbs_sheet = "📊 WBS & Effort"
    cost_start_row = 18

    ws.merge_cells(f"A{cost_start_row}:H{cost_start_row}")
    ws[f"A{cost_start_row}"] = "Cost Calculation (formulas reference Rate Card above and WBS sheet)"
    ws[f"A{cost_start_row}"].font = bold(12, "FFFFFF")
    ws[f"A{cost_start_row}"].fill = hfill(DARK_BLUE)
    ws[f"A{cost_start_row}"].alignment = center()
    ws.row_dimensions[cost_start_row].height = 24

    cost_headers = ["Epic / Module", "Dev (md)", "QA (md)", "BA (md)", "PM (md)",
                    "Dev Cost", "QA Cost", "BA Cost", "PM Cost", "Subtotal", "Buffer%", "Total"]
    cwidths = [38, 10, 10, 10, 10, 16, 16, 14, 14, 16, 10, 18]
    for col, (h, w) in enumerate(zip(cost_headers, cwidths), 1):
        header_cell(ws, cost_start_row + 1, col, h, MID_BLUE)
        ws.column_dimensions[get_column_letter(col)].width = w

    # WBS sheet row references for epics
    wbs_epic_rows = {w["id"]: (i + 2) for i, w in enumerate(WBS_ITEMS) if w["type"] == "Epic"}

    # Dev/QA/BA/PM rate rows in THIS sheet (Rate Card starts at row 3)
    rate_rows = {"Dev": 4, "QA": 7, "BA": 10, "PM": 13}  # Mid seniority rows

    data_row = cost_start_row + 2
    epics = [w for w in WBS_ITEMS if w["type"] == "Epic"]

    for epic in epics:
        wbs_row = wbs_epic_rows[epic["id"]]
        alt = LIGHT_BLUE if data_row % 2 == 0 else "FFFFFF"

        ws.cell(data_row, 1, epic["title"]).fill = hfill(alt)
        ws.cell(data_row, 1).border = thin_border()
        ws.cell(data_row, 1).font = Font(bold=True, size=10)

        # Effort columns — reference WBS sheet
        effort_cols = [(2, "E"), (3, "F"), (4, "G"), (5, "H")]
        for dest_col, wbs_col in effort_cols:
            c = ws.cell(data_row, dest_col, f"='{wbs_sheet}'!{wbs_col}{wbs_row}")
            c.number_format = "0.0"
            c.fill = hfill(alt)
            c.border = thin_border()
            c.alignment = center()

        # Cost columns — KL-P-002: formulas only
        # Dev cost = Dev_md_cell * Dev_mid_rate_cell
        rate_sheet = "💰 Rate Card & Cost"
        ws.cell(data_row, 6,  f"=B{data_row}*C{rate_rows['Dev']}").number_format = '#,##0 "USD"'
        ws.cell(data_row, 7,  f"=C{data_row}*C{rate_rows['QA']}").number_format  = '#,##0 "USD"'
        ws.cell(data_row, 8,  f"=D{data_row}*C{rate_rows['BA']}").number_format  = '#,##0 "USD"'
        ws.cell(data_row, 9,  f"=E{data_row}*C{rate_rows['PM']}").number_format  = '#,##0 "USD"'
        ws.cell(data_row, 10, f"=F{data_row}+G{data_row}+H{data_row}+I{data_row}").number_format = '#,##0 "USD"'
        ws.cell(data_row, 11, f"{epic['buffer']}%").number_format = "0\"%\""
        ws.cell(data_row, 12, f"=J{data_row}*(1+K{data_row}/100)").number_format = '#,##0 "USD"'

        for col in range(6, 13):
            ws.cell(data_row, col).fill = hfill(alt)
            ws.cell(data_row, col).border = thin_border()

        data_row += 1

    # Grand total
    first_data = cost_start_row + 2
    last_data = data_row - 1
    ws.cell(data_row, 1, "GRAND TOTAL").font = bold(11, "FFFFFF")
    ws.cell(data_row, 1).fill = hfill(DARK_BLUE)
    ws.cell(data_row, 1).border = thin_border()
    for col in range(2, 13):
        c = ws.cell(data_row, col, f"=SUM({get_column_letter(col)}{first_data}:{get_column_letter(col)}{last_data})")
        c.font = bold(11, "FFFFFF")
        c.fill = hfill(DARK_BLUE)
        c.number_format = '#,##0 "USD"'
        c.border = thin_border()
        c.alignment = center()

    ws.freeze_panes = f"A{cost_start_row + 2}"


def build_assumptions_sheet(wb: Workbook):
    ws = wb.create_sheet("📝 Assumptions & Exclusions")

    ws.merge_cells("A1:D1")
    ws["A1"] = "Technical Assumptions & Out-of-Scope Exclusions"
    ws["A1"].font = bold(14, "FFFFFF")
    ws["A1"].fill = hfill(DARK_BLUE)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    # Assumptions
    ws.merge_cells("A3:D3")
    ws["A3"] = "ASSUMPTIONS"
    ws["A3"].font = bold(12, "FFFFFF")
    ws["A3"].fill = hfill(MID_BLUE)
    ws["A3"].alignment = center()

    for col, h in enumerate(["#", "Category", "Assumption", "Impact if Wrong"], 1):
        header_cell(ws, 4, col, h, MID_BLUE)

    assumptions = [
        ("1", "Infrastructure",  "Client provides PostgreSQL 15+ with pgvector extension enabled", "Migration required if older version"),
        ("2", "LLM Contract",    "Anthropic Enterprise API contract with no-training-data clause is in place before production launch (NFR-04)", "Legal/compliance risk if standard API used"),
        ("3", "Rate Card",       "Rates in this document are indicative (Mid seniority). Final rates to be confirmed with client", "Cost delta up to ±40% depending on seniority mix"),
        ("4", "Document Quality","Input documents are reasonably structured. Machine-readable PDFs preferred; scanned PDFs may have lower analysis accuracy", "OCR errors may degrade analysis quality"),
        ("5", "Historical Data", "Effort estimates are based on AI analysis only. Historical project calibration data (FR-12) not yet available at MVP stage", "Estimates carry ±25% uncertainty without historical data"),
        ("6", "Language",        "Primary document language is English. Vietnamese language support available but may have lower LLM analysis quality", "Non-English documents may produce fewer findings"),
        ("7", "Concurrency",     "MVP targets ≤10 concurrent users. Scaling to NFR-02 (50 concurrent) requires additional worker infrastructure investment", "Additional infrastructure cost for scale-up"),
        ("8", "Integrations",    "Jira/Linear/ClickUp OAuth credentials must be configured per organization by admin before sync feature is available", "Sync feature unavailable without prior configuration"),
    ]
    for row, (num, cat, assumption, impact) in enumerate(assumptions, start=5):
        alt = LIGHT_BLUE if row % 2 == 0 else "FFFFFF"
        data_cell(ws, row, 1, num, bg=alt, align=center())
        data_cell(ws, row, 2, cat, bg=alt, bold_=True)
        data_cell(ws, row, 3, assumption, bg=alt)
        data_cell(ws, row, 4, impact, bg=YELLOW_BG)
        ws.row_dimensions[row].height = 45

    # Out of Scope
    excl_start = 5 + len(assumptions) + 2
    ws.merge_cells(f"A{excl_start}:D{excl_start}")
    ws[f"A{excl_start}"] = "OUT OF SCOPE (Auto-extracted from Risk & Gap Analysis)"
    ws[f"A{excl_start}"].font = bold(12, "FFFFFF")
    ws[f"A{excl_start}"].fill = hfill(RED)
    ws[f"A{excl_start}"].alignment = center()
    ws.row_dimensions[excl_start].height = 24

    for col, h in enumerate(["#", "Item", "Reason Out of Scope", "Source"], 1):
        header_cell(ws, excl_start + 1, col, h, RED)

    exclusions = [
        ("1", "Frontend / Web Dashboard UI",          "SRS defines backend API only. Frontend development is a separate engagement.", "Architecture Decision"),
        ("2", "Mobile application (iOS / Android)",   "SRS specifies Web Dashboard only (NFR-07). Mobile not mentioned.", "NFR-07"),
        ("3", "Delta Analysis (v1 vs v2 comparison)", "FR-06 identified as a Gap — not implemented in current sprint scope.", "GAP-04"),
        ("4", "Word (.docx) export for Clarifications","Export format not implemented. Identified as a gap.", "GAP-03"),
        ("5", "Historical Learning calibration data",  "No historical project database exists at MVP stage (FR-12 cold-start risk).", "RSK-03"),
        ("6", "Real-time WebSocket notifications",     "Current polling mechanism sufficient for MVP. WebSocket is a post-MVP enhancement.", "SUG-01"),
        ("7", "Multi-language UI (i18n)",              "UI internationalization not specified in SRS.", "Not in SRS"),
        ("8", "Third-party SSO (Google/SAML)",         "SRS specifies JWT auth only. SSO integration is a future enhancement.", "Not in SRS"),
        ("9", "Data migration from legacy systems",    "No legacy system integration specified. Historical data import is manual CSV.", "RSK-03"),
    ]
    for row, (num, item, reason, source) in enumerate(exclusions, start=excl_start + 2):
        alt = RED_BG if row % 2 == 0 else "FFFFFF"
        data_cell(ws, row, 1, num,    bg=alt, align=center())
        data_cell(ws, row, 2, item,   bg=alt, bold_=True)
        data_cell(ws, row, 3, reason, bg=alt)
        data_cell(ws, row, 4, source, bg=alt, align=center())
        ws.row_dimensions[row].height = 40

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 22


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    print("Building workbook...")
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    print("  [1/6] Cover & Summary")
    build_cover_sheet(wb)
    print("  [2/6] Analysis Results")
    build_analysis_sheet(wb)
    print("  [3/6] Test Cases")
    build_test_cases_sheet(wb)
    print("  [4/6] WBS & Effort")
    build_wbs_sheet(wb)
    print("  [5/6] Rate Card & Cost")
    build_rate_cost_sheet(wb)
    print("  [6/6] Assumptions & Exclusions")
    build_assumptions_sheet(wb)

    wb.save(OUTPUT_FILE)
    print(f"\nOK Report saved to: {OUTPUT_FILE}")
    print(f"   Sheets: {[s.title for s in wb.worksheets]}")
    print(f"\nSummary:")
    print(f"   Analysis results : {len(ANALYSIS_RESULTS)} items")
    print(f"   Test cases       : {len(TEST_CASES)} items")
    print(f"   WBS items        : {len(WBS_ITEMS)} items ({sum(1 for w in WBS_ITEMS if w['type'] == 'Task')} tasks)")


if __name__ == "__main__":
    main()
