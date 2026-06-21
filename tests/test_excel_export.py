import uuid
import io
import openpyxl
import pytest
from unittest.mock import MagicMock, patch
from decimal import Decimal
from src.models.wbs_item import WBSItem, WBSItemType, Complexity
from src.models.rate_card import RateCard, RateCardEntry
from src.estimation.excel_export import (
    build_rate_card_sheet,
    build_wbs_sheet,
    build_summary_sheet,
    Workbook,
)


def make_rate_card():
    rc = MagicMock(spec=RateCard)
    rc.id = uuid.uuid4()
    rc.currency = "USD"
    rc.name = "Standard"
    return rc


def make_entries():
    entries = []
    for role, seniority, rate in [("dev", "mid", 500), ("qa", "mid", 300), ("ba", "mid", 400), ("pm", "mid", 500)]:
        e = MagicMock(spec=RateCardEntry)
        e.role = role
        e.seniority = seniority
        e.daily_rate = Decimal(str(rate))
        entries.append(e)
    return entries


def make_wbs_items(n=3):
    items = []
    for i in range(n):
        item = MagicMock(spec=WBSItem)
        item.id = uuid.uuid4()
        item.title = f"Feature {i+1}"
        item.description = f"Description {i+1}"
        item.item_type = MagicMock()
        item.item_type.value = "story"
        item.complexity = MagicMock()
        item.complexity.value = "medium"
        item.effort_dev_md = Decimal("5")
        item.effort_qa_md = Decimal("2.5")
        item.effort_ba_md = Decimal("0.5")
        item.effort_pm_md = Decimal("0.5")
        item.risk_buffer_pct = Decimal("20")
        items.append(item)
    return items


def test_summary_sheet_uses_formulas_not_hardcoded():
    """KL-P-002: cost cells must be formulas."""
    wb = Workbook()
    wb.remove(wb.active)
    rc = make_rate_card()
    entries = make_entries()
    items = make_wbs_items(2)

    build_rate_card_sheet(wb, rc, entries)
    build_wbs_sheet(wb, items)
    ws = build_summary_sheet(wb, items, rc, entries)

    # Dev cost cell (col C, row 2) must be a formula
    dev_cost_cell = ws.cell(2, 3)
    assert str(dev_cost_cell.value).startswith("="), "Dev cost must be a formula, not hardcoded"

    # Total cost cell (col I, row 2) must be a formula
    total_cell = ws.cell(2, 9)
    assert str(total_cell.value).startswith("="), "Total cost must be a formula, not hardcoded"


def test_rate_card_sheet_has_correct_headers():
    wb = Workbook()
    wb.remove(wb.active)
    rc = make_rate_card()
    entries = make_entries()
    ws = build_rate_card_sheet(wb, rc, entries)
    assert ws.cell(1, 1).value == "Role"
    assert ws.cell(1, 2).value == "Seniority"


def test_wbs_sheet_row_count():
    wb = Workbook()
    wb.remove(wb.active)
    items = make_wbs_items(5)
    ws = build_wbs_sheet(wb, items)
    # Header + 5 rows
    assert ws.max_row == 6


def test_summary_grand_total_formula():
    wb = Workbook()
    wb.remove(wb.active)
    rc = make_rate_card()
    entries = make_entries()
    items = make_wbs_items(3)
    build_rate_card_sheet(wb, rc, entries)
    build_wbs_sheet(wb, items)
    ws = build_summary_sheet(wb, items, rc, entries)

    grand_total_row = 1 + len(items) + 1
    grand_total_cell = ws.cell(grand_total_row, 9)
    assert str(grand_total_cell.value).startswith("=SUM"), "Grand total must use SUM formula"
